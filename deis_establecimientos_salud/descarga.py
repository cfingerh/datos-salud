import json
import requests
import os
from datetime import datetime

# url = "https://deis.minsal.cl/deisajax?action=wp_ajax_ninja_tables_public_action&table_id=2736&target_action=get-all-data&default_sorting=manual_sort&skip_rows=0&limit_rows=0&chunk_number=1&ninja_table_public_nonce=ebeeef3d63"
url_buscar = "https://deis.minsal.cl/deisajax?action=wp_ajax_ninja_tables_public_action&table_id=2889&target_action=get-all-data&default_sorting=manual_sort&skip_rows=0&limit_rows=0&ninja_table_public_nonce=c79be707e6"
headers = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_10_1) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.95 Safari/537.36"}
resp = requests.get(url_buscar, headers=headers)

datas_json_descargada = json.loads(resp.text)

for data_json in datas_json_descargada:
    if data_json["value"]["ver"].find("Establecimientos DEIS MINSAL") != -1:
        url = data_json["value"]["ver"]
        break


print("La url de descarga es: ", url)


folder_descargas = os.getcwd() + "/" f"descargas_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}"
os.mkdir(folder_descargas)


archivo_xls = folder_descargas + "/datos.xlsx"
r = requests.get(url, allow_redirects=True)

open(archivo_xls, "wb").write(r.content)


import openpyxl
import hashlib

workbook = openpyxl.Workbook()
workbook = openpyxl.load_workbook(archivo_xls)


if "Establecimientos Vigentes" not in workbook.sheetnames:
    print("No se encontró la hoja 'Establecimientos Vigentes'")

sheet = workbook["Establecimientos Vigentes"]
# Se obtendrá un hash de todo el texto de la hoja de cálculo (desde la línea 2), para verificar si se ha descargado una nueva versión de la hoja de cálculo.

hash_texto = ""
for i in range(2, sheet.max_row + 1):
    for j in range(1, sheet.max_column + 1):
        hash_texto += str(sheet.cell(row=i, column=j).value)


hash_texto = hashlib.sha256(hash_texto.encode()).hexdigest()

print("El hash de la hoja de cálculo es: ", hash_texto)

{"hash": hash_texto, "filas": sheet.max_row, "columnas": sheet.max_column, "cantidad_registros": sheet.max_row - 1}


# Se guardará el hash en un archivo en la carpeta de descargas.
with open(folder_descargas + "/hash_establecimientos_vigentes.txt", "w") as f:
    f.write(hash_texto)
